"""
Excel Data Reader Utility - Reads test data from Excel files
"""
from openpyxl import load_workbook
from pathlib import Path


class ExcelDataReader:
    """Read test data from Excel files"""
    
    def __init__(self, file_path):
        """Initialize with Excel file path"""
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        self.workbook = load_workbook(self.file_path)
    
    def get_sheet_data(self, sheet_name):
        """
        Read entire sheet and return list of dictionaries
        First row treated as headers
        """
        if sheet_name not in self.workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {self.workbook.sheetnames}")
        
        sheet = self.workbook[sheet_name]
        headers = []
        data = []
        
        for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            if idx == 1:
                # First row is headers
                headers = [cell for cell in row if cell is not None]
            else:
                # Convert rows to dictionaries
                if any(row):  # Skip empty rows
                    row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                    data.append(row_dict)
        
        return data
    
    def get_single_value(self, sheet_name, row_idx, col_idx):
        """Get a single cell value"""
        sheet = self.workbook[sheet_name]
        return sheet.cell(row=row_idx, column=col_idx).value
    
    def get_all_sheets(self):
        """Get all available sheet names"""
        return self.workbook.sheetnames
    
    def close(self):
        """Close workbook"""
        self.workbook.close()
